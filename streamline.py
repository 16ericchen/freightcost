import re
import os
from os import path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings
warnings.simplefilter("ignore")
trackingDict,trackingList,changedList,projectPrice = {},{},{},{}
trackingList ={393885586820: 131.58, 393921912349: 248.81, 393923166019: 366.99, 770890530929: 125.39, 771126308493: 137.66, 771154179814: 107.42, 632466449362: 341.14, 603856722221: 423.25, 369904417946: 218.4, 393666855714: 232.0, 556726543686: 118.11}
invoiceInfo,projectList = [],[]
os.chdir(path.dirname(__file__))
def receiving(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    rows = list(sheet.iter_rows(min_row=1800, max_row=sheet.max_row))
    for row in reversed(rows):
        if row[14].value in trackingList.keys() and row[10].value:
            originalProjectName = row[10].value.strip()
            modifiedProjectName = originalProjectName.replace('\xa0','').split(' ')
            curMax = eval(str(row[9].value))
            # might need to check to see if for total quantity of a project not just that single row because the same project could have multiple parts that are in seperate rows but have the same tracking number
            if 'TO' not in modifiedProjectName:
                if originalProjectName not in changedList.keys():
                    addProjectName(modifiedProjectName,originalProjectName,row[14].value,curMax)
                else:
                    if row[14].value in trackingDict.keys():
                        final = modifiedProjectName[0][0] + re.sub('[ABCD]','',modifiedProjectName[0][1:].upper())
                        checkDict(trackingDict[row[14].value],final,curMax)
                    else:
                        trackingDict[row[14].value] = {changedList[originalProjectName] : curMax}
            else:
                if ';' in originalProjectName:
                    print(originalProjectName)
                else:
                    projectName = ' '
                    if projectName not in changedList.keys():
                        addProjectName(modifiedProjectName[2:],projectName,row[14].value,curMax)
                    else:
                        if row[14].value in trackingDict.keys():
                            final = modifiedProjectName[2][0] + re.sub('[ABCD]','',modifiedProjectName[2][1:].upper())
                            print('final',final)
                            checkDict(trackingDict[row[14].value],final,curMax)
                        else:
                            trackingDict[row[14].value] = {changedList[projectName] : curMax}
    for x in trackingDict:
        if len(trackingDict[x])>1:
            test = dict(sorted(trackingDict[x].items(), key=lambda item: item[1],reverse=True))
            checkDict(projectPrice,list(test)[0],trackingList[x])
        else:
            checkDict(projectPrice,list(trackingDict[x])[0],trackingList[x])

def modifyProjectName(originalString,stringList):
    i = 1
    final = stringList[0][0] + re.sub('[ABCD]','',stringList[0][1:].upper())
    if '' in stringList:
        stringList.remove('')
        final = stringList[2]
        i = 3
    for x in range(i,len(stringList)):
        if '-' in stringList[x]:
            test = stringList[x].split('-')
            final += ' '+test[0]
            break
        final+=" "+stringList[x]
    changedList[originalString] = final

def addProjectName(modifiedProjectName,originalProjectName,trackingNumber,curMax):
    # modify project name section
    final = modifiedProjectName[0][0] + re.sub('[ABCD]','',modifiedProjectName[0][1:].upper())
    for x in range(1,len(modifiedProjectName)):
        if '-' in modifiedProjectName[x]:
            test = modifiedProjectName[x].split('-')
            final += ' '+test[0]
            break
        final+=" "+modifiedProjectName[x]
    changedList[originalProjectName] = final
    # checking if the tracking number is in the dictionary already
    # we go by tracking number because project name is not unique in the sense that two different projects could have the same 
    if trackingNumber in trackingDict.keys():
        checkDict(trackingDict[trackingNumber],final,curMax)
    else:
        trackingDict[trackingNumber] = {final : curMax}

def checkDict(dict,projectName,value):
    if projectName in dict.keys():
        dict[projectName] += value
    else:
        dict[projectName] = value