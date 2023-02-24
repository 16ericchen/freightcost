import re
import os
from os import path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings
warnings.simplefilter("ignore")
trackingDict,trackingList,changedList,projectPrice = {},{},{},{}

invoiceInfo,projectList = [],[]
os.chdir(path.dirname(__file__))
#Column Numbers : PO# = 2,APN = 4,MFN = 5,Date = 1,Project = 10,Tracking = 14,Vendor = 12, MFG = 13, QTY = 9
# def receiving(file):
#     wb = load_workbook(filename=file)
#     sheet = wb.active
#     rows = list(sheet.iter_rows(min_row=1800, max_row=sheet.max_row))
#     for row in reversed(rows):
#         if row[14].value in trackingList.keys() and row[10].value:
#             originalProjectName = row[10].value.strip()
#             modifiedProjectName = originalProjectName.replace('\xa0','').split(' ')
#             curMax = eval(str(row[9].value))
#             # might need to check to see if for total quantity of a project not just that single row because the same project could have multiple parts that are in seperate rows but have the same tracking number
#             if 'TO' not in modifiedProjectName:
#                 if originalProjectName not in changedList.keys():
#                     # modify project name section
#                     final = modifiedProjectName[0][0] + re.sub('[ABCD]','',modifiedProjectName[0][1:].upper())
#                     for x in range(1,len(modifiedProjectName)):
#                         if modifiedProjectName == '-':
#                             break
#                         final+=" "+modifiedProjectName[x]
#                     changedList[originalProjectName] = final
#                     # checking if the tracking number is in the dictionary already
#                     # we go by tracking number because project name is not unique in the sense that two different projects could have the same 
#                     if row[14].value in trackingDict.keys():
#                         checkDict(trackingDict[row[14].value],final,curMax)
#                     else:
#                         trackingDict[row[14].value] = {final : curMax}
#                 else:
#                     if row[14].value in trackingDict.keys():
#                         checkDict(trackingDict[row[14].value],final,curMax)
#                     else:
#                         trackingDict[row[14].value] = {changedList[originalProjectName] : curMax}
#             else:
#                 if ';' in originalProjectName:
#                     print(originalProjectName)
#                 else:
#                     if '' in modifiedProjectName:
#                         modifiedProjectName.remove('')
#                     projectName = modifiedProjectName[2]
#                     for x in range(3,len(modifiedProjectName)):
#                         projectName += ' '+modifiedProjectName[x]
#                     if projectName not in changedList.keys()
#     for x in trackingDict:
#         if len(trackingDict[x])>1:
#             test = dict(sorted(trackingDict[x].items(), key=lambda item: item[1],reverse=True))
#             checkDict(projectPrice,list(test)[0],trackingList[x])
#         else:
#             checkDict(projectPrice,list(trackingDict[x])[0],trackingList[x])
                



# def receiving(file):
#     wb = load_workbook(filename=file)
#     sheet = wb.active
#     print(trackingList)
#     rows = list(sheet.iter_rows(min_row=1800, max_row=sheet.max_row))
#     for row in reversed(rows):
#         if row[14].value in trackingList.keys() and row[10].value:
#             originalProjectName = row[10].value.strip()
#             print(originalProjectName)
#             modifiedProjectName = originalProjectName.replace('\xa0','').split(' ')
#             curMax = eval(str(row[9].value))
#             # might need to check to see if for total quantity of a project not just that single row because the same project could have multiple parts that are in seperate rows but have the same tracking number
#             if 'TO' not in modifiedProjectName:
#                 if originalProjectName not in changedList.keys():
#                     addProjectName(modifiedProjectName,originalProjectName,row[14].value,curMax)
#                 else:
#                     if row[14].value in trackingDict.keys():
#                         final = modifiedProjectName[0][0] + re.sub('[ABCD]','',modifiedProjectName[0][1:].upper())
#                         checkDict(trackingDict[row[14].value],final,curMax)
#                     else:
#                         trackingDict[row[14].value] = {changedList[originalProjectName] : curMax}
#             else:
#                 if ';' in originalProjectName:
#                     print(originalProjectName)
#                 else:
#                     if '' in modifiedProjectName:
#                         modifiedProjectName.remove('')
#                     projectName = modifiedProjectName[2]
#                     for x in range(3,len(modifiedProjectName)):
#                         projectName += ' '+modifiedProjectName[x]
#                     if projectName not in changedList.keys():
#                         addProjectName(modifiedProjectName[2:],projectName,row[14].value,curMax)
#                     else:
#                         if row[14].value in trackingDict.keys():
#                             final = modifiedProjectName[2][0] + re.sub('[ABCD]','',modifiedProjectName[2][1:].upper())
#                             print('final',final)
#                             for x in range(1,len(modifiedProjectName)):
#                                 if '-' in modifiedProjectName[x]:
#                                     test = modifiedProjectName[x].split('-')
#                                     final += ' '+test[0]
#                                     break
#                                 final+=" "+modifiedProjectName[x]
#                             print(projectName)
#                             print('final after ',final)
#                             checkDict(trackingDict[row[14].value],final,curMax)
#                         else:
#                             trackingDict[row[14].value] = {changedList[projectName] : curMax}
#     for x in trackingDict:
#         if len(trackingDict[x])>1:
#             test = dict(sorted(trackingDict[x].items(), key=lambda item: item[1],reverse=True))
#             checkDict(projectPrice,list(test)[0],trackingList[x])
#         else:
#             checkDict(projectPrice,list(trackingDict[x])[0],trackingList[x])


# def addProjectName(modifiedProjectName,originalProjectName,trackingNumber,curMax):
#     # modify project name section
#     final = modifiedProjectName[0][0] + re.sub('[ABCD]','',modifiedProjectName[0][1:].upper())
#     for x in range(1,len(modifiedProjectName)):
#         if '-' in modifiedProjectName[x]:
#             test = modifiedProjectName[x].split('-')
#             final += ' '+test[0]
#             break
#         final+=" "+modifiedProjectName[x]
#     changedList[originalProjectName] = final
#     # checking if the tracking number is in the dictionary already
#     # we go by tracking number because project name is not unique in the sense that two different projects could have the same 
#     if trackingNumber in trackingDict.keys():
#         checkDict(trackingDict[trackingNumber],final,curMax)
#     else:
#         trackingDict[trackingNumber] = {final : curMax}

# def checkDict(dict,projectName,value):
#     if projectName in dict.keys():
#         dict[projectName] += value
#     else:
#         dict[projectName] = value  





# def receiving(file):
#     wb = load_workbook(filename=file)
#     sheet = wb.active
#     rows = list(sheet.iter_rows(min_row=1800, max_row=sheet.max_row))
#     for row in reversed(rows):
#         if row[14].value in trackingList.keys() and row[10].value:
#             originalProjectName = row[10].value.strip()
#             modifiedProjectName = originalProjectName.replace('\xa0','').split(' ')
#             curMax = eval(str(row[9].value))
#             # might need to check to see if for total quantity of a project not just that single row because the same project could have multiple parts that are in seperate rows but have the same tracking number
#             if 'TO' not in modifiedProjectName:
#                 if originalProjectName not in changedList.keys():
#                     addProjectName(modifiedProjectName,originalProjectName,row[14].value,curMax)
#                 else:
#                     if row[14].value in trackingDict.keys():
#                         final = modifiedProjectName[0][0] + re.sub('[ABCD]','',modifiedProjectName[0][1:].upper())
#                         checkDict(trackingDict[row[14].value],final,curMax)
#                     else:
#                         trackingDict[row[14].value] = {changedList[originalProjectName] : curMax}
#             else:
#                 if ';' in originalProjectName:
#                     print(originalProjectName)
#                 else:
#                     projectName = ' '
#                     if projectName not in changedList.keys():
#                         addProjectName(modifiedProjectName[2:],projectName,row[14].value,curMax)
#                     else:
#                         if row[14].value in trackingDict.keys():
#                             final = modifiedProjectName[2][0] + re.sub('[ABCD]','',modifiedProjectName[2][1:].upper())
#                             print('final',final)
#                             checkDict(trackingDict[row[14].value],final,curMax)
#                         else:
#                             trackingDict[row[14].value] = {changedList[projectName] : curMax}
#     for x in trackingDict:
#         if len(trackingDict[x])>1:
#             test = dict(sorted(trackingDict[x].items(), key=lambda item: item[1],reverse=True))
#             checkDict(projectPrice,list(test)[0],trackingList[x])
#         else:
#             checkDict(projectPrice,list(trackingDict[x])[0],trackingList[x])

# def modifyProjectName(originalString,stringList):
#     i = 1
#     final = stringList[0][0] + re.sub('[ABCD]','',stringList[0][1:].upper())
#     if '' in stringList:
#         stringList.remove('')
#         final = stringList[2]
#         i = 3
#     for x in range(i,len(stringList)):
#         if '-' in stringList[x]:
#             test = stringList[x].split('-')
#             final += ' '+test[0]
#             break
#         final+=" "+stringList[x]
#     changedList[originalString] = final

# def addProjectName(modifiedProjectName,originalProjectName,trackingNumber,curMax):
#     # modify project name section
#     final = modifiedProjectName[0][0] + re.sub('[ABCD]','',modifiedProjectName[0][1:].upper())
#     for x in range(1,len(modifiedProjectName)):
#         if '-' in modifiedProjectName[x]:
#             test = modifiedProjectName[x].split('-')
#             final += ' '+test[0]
#             break
#         final+=" "+modifiedProjectName[x]
#     changedList[originalProjectName] = final
#     # checking if the tracking number is in the dictionary already
#     # we go by tracking number because project name is not unique in the sense that two different projects could have the same 
#     if trackingNumber in trackingDict.keys():
#         checkDict(trackingDict[trackingNumber],final,curMax)
#     else:
#         trackingDict[trackingNumber] = {final : curMax}

# def checkDict(dict,projectName,value):
#     if projectName in dict.keys():
#         dict[projectName] += value
#     else:
#         dict[projectName] = value  








# Coulmn Number: Tracking = 9, cost = 11, invoice number = 4,invoice date = 3,total amount = 6
def invoice(file):
    wb = load_workbook(filename=file)
    sheet = wb.active
    # [posting date,invoice number,invoice date,invoice amount]
    invoiceInfo.extend([datetime.now().strftime('%m/%d/%Y'),sheet.cell(row=2,column=4).value,datetime.strptime(sheet.cell(row=2,column=3).value, '%Y%m%d').strftime('%m/%d/%Y'),sheet.cell(row=2,column=6).value])
    for row in sheet.iter_rows(min_row=2, min_col=1,max_col=12):
        if row[9].value in trackingList.keys() and row[11].value != '' :
            trackingList[row[9].value]+=float(row[11].value)
        if row[9].value not in trackingList.keys():
            trackingList[row[9].value] = float(row[11].value)

def freight(file):
    # print(trackingDict)
    # for x in trackingDict:
    #     for y in trackingDict[x]:
    #         print(y)
    # add invoice information to the freight cost document
    # col 2 = posting date(the date the program is being run) in format MM/DD/YYYY# col 3 = Fedex Invoice Number(found in invoice excel sheet)# col 4 = Invoice Date(Found in the invoice excel sheet) # col 5 = Invoice Amount(Found in the invoice excel sheet)
    wb = load_workbook(filename=file)
    sheet = wb[wb.sheetnames[-1]]
    nextRow = 0
    for row in sheet.iter_rows(min_row=1,min_col=4):
    # this condition is to fin the last row in the table ie the sum row
        if "=" in str(row[1].value):
            endRow = row[1].row
    # This iteration is to get all the project names in one row so
        if row[1].row == 1:
            for x in row:
                if x.value:
                    projectList.append(x.value)
    # this is to find the next empty row so to put the information in
        if row[1].value == None and nextRow == 0:
            nextRow = row[1].row
    my_list = [elem[0] for elem in trackingDict.values()]
    my_list = sorted(my_list)
    counter = 0
    pointer = 2
    for x in range(len(invoiceInfo)):
        print(nextRow,x+1,invoiceInfo[x])
    for x in my_list:
        if x in projectList:
            # dont update price yet need to get total list of projects then insert row by row and update according to invoice number.
            # nvm since we are going based off of one invoice
            # so in the future if they want to load in multiple invoices we just run the functions multiple times
            # this can be continue or store the values that are already in the projectlist in another list and search for them later when inputting the price into the row
            # we can update the price here because the column will be inserted after this row
            print('Update price')
        else:
            for i in range(pointer,len(projectList)):
                if x<projectList[i]:
                    # need to add the project name to row 1 and need to add the total to the last row
                    #  as well as style them so that the fill color is green and there are borders around each cell
                    # insert column before column number(counter+i)
                    # update price in the column
                    # or i can store the column number with the price 
                    pointer = i
                    # print(x,projectList[i])
                    # print(counter+i)
                    # sheet.insert_cols(counter+i)
                    counter += 1
                    break
    
    # this is the formula that needs to be in the last column
    print('=SUM(F'+str(endRow)+':'+get_column_letter(len(projectList)-5)+str(endRow)+')')
    # sheet.cell(row=endRow,column = len(projectList)+3,value = '=SUM(F'+str(endRow)+':'+get_column_letter(len(projectList)-5)+str(endRow)+')')
    # (ws.cell(row=row_number, column=column_number).value)
    wb.save(filename=file)

if __name__ == '__main__':
    # receivingFolder = r"N:\WHS\Receiving Log"
    receivingFolder = r"C:\Users\EricChen\Desktop\Python Scripts\In Progress\Freight Cost\New folder"
    receivingLog = ''
    freightFile = r"C:\Users\EricChen\Desktop\Python Scripts\In Progress\Freight Cost\New folder (2)\2022 Freight Cost R3.xlsx"
    for y in os.listdir(receivingFolder):
        if y.endswith(('.xlsx')):
            receivingLog = receivingFolder+'\\'+y
    if receivingFolder == '':
        print("Receiving Log Not Found")
        print("Please Double Check File and Restart Program")
        input()
    for x in os.listdir(os.getcwd()):
        if x.endswith(('.xlsx')):
            invoice(x)
    receiving(receivingLog)
    freight(freightFile)